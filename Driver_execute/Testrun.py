# coding=utf-8
'''
@Time   :2020/12/15 18:20
@Author :六月
@Email  :juneren26@gmail.com
@File   :Testrun.py
@IDE    :PyCharm
'''

import openpyxl
from openpyxl.styles import PatternFill, Font
from FSkeys.keywords import Webkeys
from Logs.log import Logger
# 日志
log = Logger().get_logger()
# 访问excel内容
excel = openpyxl.load_workbook('../Testcase/FS_login.xlsx')
sheet = excel['Sheet1']
log.info('获取{}内容成功,现在开始执行自动化测试......'.format(sheet))
# 读取excel内容，实现文件驱动自动化执行
for value in sheet.values:
    # 定义一个字典接收excel中的所有参数
    args = {}
    args['name'] = value[2]
    args['value'] = value[3]
    args['txt'] = value[4]
    args['expect'] = value[6]
    # 基于A列判断是否为测试用例的数据
    if type(value[0]) is int:
        '''
        读取关键字过程中需要区分：
            关键字驱动类的实例化
            断言类型的关键字
        '''
        # 判断是否实例化
        if value[1] == 'browser':
            log.info('现在执行关键:{},操作描述{}'.format(value[1],value[5]))
            wk = Webkeys(value[4])
        # 断言关键字执行结果
        elif 'assert' in value[1]:
            log.info('现在执行关键:{},操作描述{}'.format(value[1], value[5]))
            status = getattr(wk, value[1])(**args)
            print(status)
            # 写入excel实际结果
            if status:
                sheet.cell(row=value[0] + 1, column=8).value = 'Pass'
                sheet.cell(row=value[0] + 1, column=8).fill = PatternFill('solid', fgColor='AACF91')
                sheet.cell(row=value[0] + 1, column=8).font = Font(bold=True)
            else:
                sheet.cell(row=value[0] + 1, column=8).value = 'Failed'
                sheet.cell(row=value[0] + 1, column=8).fill = PatternFill('solid', fgColor='FF0000')
                sheet.cell(row=value[0] + 1, column=8).font = Font(bold=True)
            excel.save('../Testcase/FS_login.xlsx')
        else:
            log.info('现在执行关键:{},操作描述{}'.format(value[1], value[5]))
            getattr(wk, value[1])(**args)
excel.close()
